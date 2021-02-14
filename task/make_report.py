import openpyxl
from openpyxl.utils import get_column_letter
import datetime
from django.conf import settings
from .models import License, ReportHistory
import os


def str_to_date(string: str, format_in='%d.%m.%Y', format_out='%Y-%m-%d') -> datetime.date:
    """
    Returns datetime.date for input string
    :param string: String like %d.%m.%Y
    :param format_in: format of input string
    :param format_out: format of output string
    :return: datetime.date object
    """
    if string is None:
        return None
    else:
        return datetime.datetime.strptime(datetime.datetime.strptime(string, format_in).strftime(format_out), format_out).date()


def date_to_str(date: datetime.date, format_out='%Y-%m-%d') -> str:
    """
    Returns str for input datetime.date
    :param date: datetime.time object
    :param format_out: format of output string
    :return: String like
    """
    if date is None:
        return None
    else:
        return date.strftime(format_out)


def change_strdate_format(string, format_in='%d.%m.%Y', format_out='%Y-%m-%d'):
    return datetime.datetime.strptime(datetime.datetime.strptime(string, format_in).strftime(format_out), format_out).strftime(format_out)


def upload_xlsx(file, file_date):

    wb = openpyxl.load_workbook(file, read_only=True)
    ws = wb.active

    report_list = []

    for row in ws.iter_rows(min_row=7):
        home_address = row[12].value
        company = row[9].value
        if row[15].value is not None:
            date_start = str_to_date(row[15].value)
        else:
            date_start = None
        if row[16].value is not None:
            date_end = str_to_date(row[16].value)
        else:
            date_end = None
        reason_to_exclude = row[18].value
        inn = row[10].value
        date_changed = str_to_date(file_date)
        str_date_changed = date_to_str(date_changed)

        if home_address is None:  # пропускать в файле строки, где нет адреса дома
            continue

        report_obj = {}

        try:
            obj = License.objects.get(home_address=home_address)  # обьект из БД с таким же дом. адресом, как в файле

            if company == obj.company and date_start == obj.date_start and date_end:  # окончание
                msg = {'event': 'Окончание', 'home_address': home_address, 'prev_company': obj.company, 'prev_inn': obj.inn,
                       'prev_date_start': date_to_str(obj.date_start), 'prev_date_end': date_to_str(obj.date_end),
                       'prev_reason_to_exclude': obj.reason_to_exclude, 'new_company': '', 'new_inn': '', 'new_date_start': None,
                       'date_changed': str_date_changed}
                obj.date_start = date_start
                obj.date_end = date_end
                obj.reason_to_exclude = reason_to_exclude
                obj.inn = inn
                obj.date_changed = date_changed
                # obj.save()
                report_obj = msg
            elif company != obj.company and date_start > obj.date_start and date_end is None and obj.date_end is None:  # новый
                msg = {'event': 'Новый', 'home_address': home_address, 'prev_company': obj.company,
                       'prev_inn': obj.inn, 'prev_date_start': date_to_str(obj.date_start), 'prev_date_end': None,
                       'prev_reason_to_exclude': obj.reason_to_exclude, 'new_company': company,
                       'new_inn': inn, 'new_date_start': date_to_str(date_start), 'date_changed': str_date_changed}
                obj.company = company
                obj.date_start = date_start
                obj.date_end = date_end
                obj.reason_to_exclude = reason_to_exclude
                obj.inn = inn
                obj.date_changed = date_changed
                # obj.save()
                try:
                    type_of_report = report_obj['event']
                    if type_of_report == 'Окончание':
                        report_obj = msg
                        report_obj['event'] = 'Переход'
                except KeyError:
                    report_obj = msg
        except License.DoesNotExist:
            msg = {'event': 'Новый', 'home_address': home_address, 'prev_company': '', 'prev_inn': '',
                   'prev_date_start': None, 'prev_date_end': None,
                   'prev_reason_to_exclude': '', 'new_company': company, 'new_inn': inn,
                   'new_date_start': date_to_str(date_start), 'date_changed': str_date_changed}
            obj = License(home_address=home_address, company=company, date_start=date_start,
                          date_end=date_end, reason_to_exclude=reason_to_exclude, inn=inn, date_changed=date_changed)
            # obj.save()
            report_obj = msg
        if report_obj:
            report_list.append(report_obj)

    wb.close()

    return report_list


def update_report_history(list_of_objs):

    for obj in list_of_objs:
        print(obj)
        new_entry = ReportHistory(event=obj['event'], home_address=obj['home_address'], prev_company=obj['prev_company'],
                                  prev_inn=obj['prev_inn'], prev_date_start=obj['prev_date_start'],
                                  prev_date_end=obj['prev_date_end'], prev_reason_to_exclude=obj['prev_reason_to_exclude'],
                                  new_company=obj['new_company'], new_inn=obj['new_inn'], new_date_start=obj['new_date_start'],
                                  date_changed=obj['date_changed'])
        new_entry.save()


def make_xlsx(date_range):
    """

    :param date_range: list of date range
    :return:
    """

    if len(date_range) == 1:
        date_range = date_range[0]
        date_name = date_to_str(date_range)
        list_of_objs = ReportHistory.objects.filter(date_changed=date_range)
    elif len(date_range) == 2:
        date_name = '--'.join([date_to_str(obj) for obj in date_range])
        list_of_objs = ReportHistory.objects.filter(date_changed__range=date_range)
    else:
        return False

    name = f'reports/Отчет за {date_name}.xlsx'
    filename = os.path.join(settings.MEDIA_ROOT, name)
    if os.path.isfile(filename):
        return filename


    wb = openpyxl.Workbook()
    ws = wb.active
    column_widths = [12, 100, 150, 11, 12, 12, 40, 100, 11, 12]
    ws['A1'] = 'Событие'
    ws['B1'] = 'Адрес дома'
    ws['C1'] = 'Предыдущая компания'
    ws['H1'] = 'Новая компания'
    ws['C2'] = 'Наименование'
    ws['D2'] = 'Инн'
    ws['E2'] = 'Дата начала полномочий'
    ws['F2'] = 'Дата окончания полномочий'
    ws['G2'] = 'Основание заключения'
    ws['H2'] = 'Наименование'
    ws['I2'] = 'Инн'
    ws['J2'] = 'Дата начала полномочий'
    for obj in list_of_objs:
        excel_row = (obj.event, obj.home_address, obj.prev_company, obj.prev_inn, obj.prev_date_start,
                     obj.prev_date_end, obj.prev_reason_to_exclude, obj.new_company, obj.new_inn,
                     obj.new_date_start)
        ws.append(excel_row)
    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = column_width

    wb.save(filename)
    return filename

